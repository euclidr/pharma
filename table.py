# -*- coding: utf-8 -*-
"""
解析出年报中的表格
"""

import re

from pprint import pprint
from typing import List

from submission_text import SubmissionText
from bs4 import BeautifulSoup, NavigableString
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


def text_from_spans(spans):
    text = ''
    for span in spans:
        text += span.get_text()
    return text


def extract_tables(html_text: str) -> List['Table']:
    """从 html 文本中提取表格"""

    soup = BeautifulSoup(html_text, 'lxml')
    html_tables = soup.find_all('table')

    items = []
    for html_table in html_tables:
        """filter tables that has title"""
        title = ''
        for parent in html_table.parents:
            if parent.name != 'div':
                break

            for prev in parent.previous_siblings:
                if isinstance(prev, NavigableString):
                    continue

                if prev.name == 'div':
                    title = prev.get_text()
                    if not title:
                        continue
                    # if 'Net Product Sales of Regeneron-Discovered' in html_table.get_text():
                    #     print('******', title, "****")

                elif prev.name == 'span':
                    title = text_from_spans(list(parent.previous_siblings))
                elif prev.name == 'p':
                    title = prev.get_text()
                    if not title.strip() or title == '&nbsp;':
                        title = ''
                        continue

                break

            if title:
                items.append({
                    'table': html_table,
                    'title': title
                })
                break

    tables = []
    for item in items:
        table = Table(item['title'], item['table'])
        if table.is_valid:
            tables.append(table)

    return tables


class LogicCell(object):
    """逻辑表格单元，一个逻辑表格单元可以包含多个表格单元，主要为了表示一个逻辑列包含了多少小列"""
    def __init__(self, start: int, end: int, text: str):
        self.start = start
        self.end = end
        self.text = text.strip()
        self.texts = [self.text]

    def __repr__(self):
        return '%s %s %s' % (
            self.start,
            self.end,
            self.text
        )

    @property
    def full_text(self):
        return '.'.join(self.texts)

    def is_same_logic_col(self, cell):
        return self.start <= cell.start \
               and self.end >= cell.end

    def merge(self, cell):
        self.texts.append(cell.text)


class TableSchema(object):
    """表格式"""

    def __init__(self, header_trs):
        self.header_trs = [tr for tr in header_trs if not isinstance(tr, NavigableString)]
        self.is_valid = False
        self.definitions = []  # type: List[LocigCell]

        self.parse()

    @property
    def col_count(self):
        return len(self.definitions)

    def get_effetive_trs(self):
        if not self.header_trs:
            return []

        first_tr = self.header_trs[0]
        tds = first_tr.find_all('td')
        if len(tds) == 1:
            return self.header_trs[2:]
        else:
            return self.header_trs

    def parse(self):
        """解析表头"""

        effective_trs = self.get_effetive_trs()

        rows = []
        for tr in effective_trs:
            row, pos = [], 0
            for td in tr.find_all('td'):
                span = int(td.get('colspan', '1'))
                cell = LogicCell(pos, pos+span, td.text)
                row.append(cell)
                pos += span
            rows.append(row)

        if not rows:
            return

        rows.sort(key=lambda x: len(x), reverse=True)

        definitions = rows[0]
        for row in rows[1:]:
            for cell in row:
                for definition in definitions:
                    if cell.is_same_logic_col(definition):
                        definition.merge(cell)

        # 取消空的列（除了第一列）
        tail = definitions[1:]
        tail = [t for t in tail if t.text.strip()]
        definitions = definitions[:1] + tail

        self.definitions = definitions

        self.is_valid = True


    def get_cell_index(self, cell: LogicCell):
        assert self.is_valid

        for idx, definition in enumerate(self.definitions):
            if definition.is_same_logic_col(cell):
                return idx

        # panic


class Table(object):

    def __init__(self, title, html_table):
        self.title = title
        self.html_table = html_table
        self.is_valid = False
        self.schema = None  # Type: TableSchema
        self.rows = []

        self.parse()

    def all_text(self):
        return self.html_table.get_text()

    def parse(self):
        first_data_tr = self.find_first_data_tr()
        if not first_data_tr:
            return

        self.remove_sup()

        header_trs = list(first_data_tr.previous_siblings)
        header_trs.reverse()
        schema = TableSchema(header_trs)
        if not schema.is_valid:
            return

        self.schema = schema

        data_trs = [first_data_tr] + list(first_data_tr.next_siblings)
        rows = []
        for data_tr in data_trs:
            row = self.parse_row(data_tr)
            if row:
                rows.append(row)

        self.rows = rows

        self.is_valid = True

    def remove_sup(self):
        for sup in self.html_table.find_all('sup'):
            sup.extract()

    def parse_row(self, data_tr):
        if isinstance(data_tr, NavigableString):
            return None

        row = [''] * self.schema.col_count
        pos = 0
        for td in data_tr.find_all('td'):
            span = int(td.get('colspan', '1'))
            item = LogicCell(pos, pos+span, td.text)
            # print(item)

            idx = self.schema.get_cell_index(item)
            # print('xxxx', col_idx)
            if idx is not None:
                row[idx] += ' ' + item.text

            pos += span

        row = [r.strip() for r in row]
        return row

    def find_first_data_tr(self):
        styles = ['background-color:#cceeff', 'background-color:#d8d8d8']
        def fit_style(val):
            if not val:
                return False

            for style in styles:
                if style in val:
                    return True

        td = self.html_table.find('td', style=fit_style)
        if td:
            return td.parent

        td = self.html_table.find('td', {'bgcolor': '#CFF0FC'})
        if td:
            return td.parent

    def print(self):
        assert self.is_valid

        header = [g.full_text for g in self.schema.definitions]
        print('\t\t'.join(header))

        for data in self.rows:
            print('\t\t'.join(data))

    def peek(self):
        print('title: ', self.title)
        print('\t\t'.join([d.full_text for d in self.schema.definitions]))
        for row in self.rows[:4]:
            print('\t\t'.join(row))

    def extract_data(self, pattern) -> dict:
        data = {}
        for idx, definition in enumerate(self.schema.definitions):
            text = definition.full_text.strip()
            m = re.search(pattern, text)
            if not m:
                continue

            col_name = m.group(1)
            for row in self.rows:
                row_name = row[0].strip()
                val = row[idx].strip()
                if not row_name or not val:
                    continue

                key = '%s.%s' % (col_name, row_name)
                data[key] = val

        return data


if __name__ == '__main__':
    parser = SubmissionTextParser('reports/Merck/0000310158/10-k/0000310158-19-000014.txt')
    file_10k = parser.get_file_10k()
    tables = extract_tables(file_10k.text)
    for table in tables:
        if table.is_valid:
            table.print()
    # print(extractor.html_text[:20])
    wb = Workbook()
    for idx, table in enumerate(tables):
        sheet_name = 'sheet_{:03d}'.format(idx)
        ws = wb.create_sheet(sheet_name)
        ws.merge_cells(
            start_row=1,
            end_row=1,
            start_column=1,
            end_column=table.schema.col_count)

        ws.cell(row=1, column=1, value=table.title)

        for c, item in enumerate(table.schema.cols):
            text = '.'.join(item.texts)
            ws.cell(row=3, column=c+1, value=text)

        for r in range(len(table.rows)):
            row = table.rows[r]
            for c in range(len(row)):
                ws.cell(row=r+4, column=c+1, value=row[c])

        # adjust table cell
        for xs_col in ws.columns:
            max_length = 0
            col_name = get_column_letter(xs_col[0].column)
            for cell in xs_col:
                if cell.coordinate in ws.merged_cells: # not check merge_cells
                    continue
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            # print(col_name)
            ws.column_dimensions[col_name].width = adjusted_width

    wb.save('merck.xlsx')


