# -*- coding: utf-8 -*-
"""
将获得的 JSON 数据转换成 Excel
"""

import config
import json


from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

data = {}

with open('result.json') as f:
    text = f.read()
    data = json.loads(text)

def get_conf_of_company(name):
    for company_conf in config.COMPANIES:
        if company_conf['name'] == name:
            return company_conf

for name, stat in data.items():
    conf = get_conf_of_company(name)
    for key, val in stat.items():
        year, _ = key.split('.')
        if int(year) in conf.get('year_in_thousand', []):
            data[name][key] = str(float(val) / 1000)

drugs = set()
for company, stat in data.items():
    for key, val in stat.items():
        _, drug_name = key.split('.')
        drugs.add('%s.%s' % (company, drug_name))

drugs = list(drugs)
drugs.sort()

years = list(range(2011, 2020))
years = [str(y) for y in years]
years.reverse()

row = 1
for idx, year in enumerate(years):
    ws.cell(row, idx+3, year)

for drug in drugs:
    row += 1
    company, drug_name = drug.split('.')
    ws.cell(row, 1, company)
    ws.cell(row, 2, drug_name)
    for idx, year in enumerate(years):
        val = data[company].get('%s.%s' % (year, drug_name), '')
        ws.cell(row, idx+3, val)

wb.save('result.xlsx')